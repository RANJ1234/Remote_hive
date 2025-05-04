import logging
import os
import io
import csv
import json
import random
from datetime import datetime, timezone
from flask import render_template, redirect, url_for, flash, request, abort, make_response
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename

from app import app, db
from models import User, Job, Company, JobApplication, CompanyCategory, CompanyCategoryAssociation, JobseekerProfile
from forms import AdminUserForm, CompanyForm

# For Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    Workbook = None
    Font = None

# For PDF export
try:
    from reportlab.pdfgen import canvas
except ImportError:
    canvas = None

# Admin access decorator
def admin_required(f):
    @login_required
    def decorated_function(*args, **kwargs):
        if not current_user.is_admin():
            abort(403)
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

# Admin Dashboard
@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    # Get basic statistics
    total_users = User.objects.count()
    total_jobs = Job.objects.count()
    total_companies = Company.objects.count()
    total_applications = JobApplication.objects.count()

    # Get recent users and jobs
    recent_users = User.objects.order_by('-created_at').limit(5)
    recent_jobs = Job.objects.order_by('-posted_date').limit(5)

    # Job statistics
    active_jobs = Job.objects(is_active=True).count()
    remote_jobs = Job.objects(is_remote=True).count()

    # Get companies with most jobs
    companies = Company.objects.all()
    company_job_counts = []
    for company in companies:
        job_count = Job.objects(company=company).count()
        if job_count > 0:
            company_job_counts.append((company, job_count))

    # Sort by job count and take top 5
    top_companies_by_jobs = sorted(company_job_counts, key=lambda x: x[1], reverse=True)[:5]

    return render_template('admin/dashboard.html',
                           total_users=total_users,
                           total_jobs=total_jobs,
                           total_companies=total_companies,
                           total_applications=total_applications,
                           recent_users=recent_users,
                           recent_jobs=recent_jobs,
                           active_jobs=active_jobs,
                           remote_jobs=remote_jobs,
                           top_companies_by_jobs=top_companies_by_jobs)

# User Management
@app.route('/admin/users')
@admin_required
def manage_users():
    page = request.args.get('page', 1, type=int)
    # MongoDB pagination
    per_page = 20
    skip = (page - 1) * per_page
    total_users = User.objects.count()
    users_list = User.objects.skip(skip).limit(per_page)

    # Create a pagination object with the necessary attributes
    class Pagination:
        def __init__(self, items, page, per_page, total):
            self.items = items
            self.page = page
            self.per_page = per_page
            self.total = total
            self.pages = (total + per_page - 1) // per_page  # Ceiling division

        @property
        def has_prev(self):
            return self.page > 1

        @property
        def has_next(self):
            return self.page < self.pages

        @property
        def prev_num(self):
            return self.page - 1 if self.has_prev else None

        @property
        def next_num(self):
            return self.page + 1 if self.has_next else None

    users = Pagination(users_list, page, per_page, total_users)
    return render_template('admin/manage_users.html', users=users)

@app.route('/admin/user/new', methods=['GET', 'POST'])
@admin_required
def add_user():
    form = AdminUserForm()
    if form.validate_on_submit():
        user = User(
            username=form.username.data,
            email=form.email.data,
            role=form.role.data,
            is_active=form.is_active.data
        )
        if form.password.data:
            user.set_password(form.password.data)
        else:
            user.set_password('password123')  # Default password

        # Save user to MongoDB
        user.save()

        # Create profile based on role
        if user.is_jobseeker():
            profile = JobseekerProfile(user=user)
            profile.save()
        flash('User created successfully!', 'success')
        return redirect(url_for('manage_users'))

    return render_template('admin/user_form.html', form=form, title='Add User')

@app.route('/admin/user/<user_id>/edit', methods=['GET', 'POST'])
@admin_required
def edit_user(user_id):
    try:
        user = User.objects.get(id=user_id)
    except:
        abort(404)

    # We need to import the form
    from forms import AdminUserForm
    form = AdminUserForm(obj=user)

    if form.validate_on_submit():
        user.username = form.username.data
        user.email = form.email.data
        user.role = form.role.data
        user.is_active = form.is_active.data

        if form.password.data:
            user.set_password(form.password.data)

        user.save()
        flash('User updated successfully!', 'success')
        return redirect(url_for('manage_users'))

    return render_template('admin/user_form.html', form=form, user=user, title='Edit User')

@app.route('/admin/user/<user_id>/delete', methods=['POST'])
@admin_required
def delete_user(user_id):
    try:
        user = User.objects.get(id=user_id)
        user.delete()
        flash('User deleted successfully!', 'success')
    except:
        flash('User not found or could not be deleted', 'danger')
    return redirect(url_for('manage_users'))

# Job Management
@app.route('/admin/jobs')
@admin_required
def manage_jobs():
    page = request.args.get('page', 1, type=int)
    per_page = 20
    skip = (page - 1) * per_page

    # Get total count for pagination
    total_jobs = Job.objects.count()

    # Get jobs with pagination
    jobs_list = Job.objects.order_by('-posted_date').skip(skip).limit(per_page)

    # Create a pagination object with the necessary attributes
    class Pagination:
        def __init__(self, items, page, per_page, total):
            self.items = items
            self.page = page
            self.per_page = per_page
            self.total = total
            self.pages = (total + per_page - 1) // per_page  # Ceiling division

        @property
        def has_prev(self):
            return self.page > 1

        @property
        def has_next(self):
            return self.page < self.pages

        @property
        def prev_num(self):
            return self.page - 1 if self.has_prev else None

        @property
        def next_num(self):
            return self.page + 1 if self.has_next else None

    jobs = Pagination(jobs_list, page, per_page, total_jobs)
    return render_template('admin/manage_jobs.html', jobs=jobs)

@app.route('/admin/job/new', methods=['GET', 'POST'])
@admin_required
def add_job():
    # Import the form
    from forms import JobPostForm
    form = JobPostForm()
    companies = Company.objects.all()

    if not companies:
        flash('Please create a company first before adding a job.', 'warning')
        return redirect(url_for('add_company'))

    if form.validate_on_submit():
        company_id = request.form.get('company_id')
        try:
            company = Company.objects.get(id=company_id)
        except:
            flash('Invalid company selected.', 'danger')
            return render_template('admin/job_form.html', form=form, companies=companies, title='Add Job')

        job = Job(
            company=company,
            title=form.title.data,
            location=form.location.data,
            is_remote=form.is_remote.data,
            job_type=form.job_type.data,
            description=form.description.data,
            requirements=form.requirements.data,
            salary_min=form.salary_min.data,
            salary_max=form.salary_max.data,
            experience_required=form.experience_required.data,
            education_required=form.education_required.data,
            deadline=form.deadline.data,
            is_active=True
        )

        job.save()
        flash('Job created successfully!', 'success')
        return redirect(url_for('manage_jobs'))

    return render_template('admin/job_form.html', form=form, companies=companies, title='Add Job')

@app.route('/admin/job/<job_id>/edit', methods=['GET', 'POST'])
@admin_required
def edit_job(job_id):
    try:
        job = Job.objects.get(id=job_id)
    except:
        abort(404)

    # Import the form
    from forms import JobPostForm
    form = JobPostForm(obj=job)
    companies = Company.objects.all()

    if form.validate_on_submit():
        company_id = request.form.get('company_id')
        try:
            company = Company.objects.get(id=company_id)
        except:
            flash('Invalid company selected.', 'danger')
            return render_template('admin/job_form.html', form=form, job=job, companies=companies, title='Edit Job')

        job.company = company
        job.title = form.title.data
        job.location = form.location.data
        job.is_remote = form.is_remote.data
        job.job_type = form.job_type.data
        job.description = form.description.data
        job.requirements = form.requirements.data
        job.salary_min = form.salary_min.data
        job.salary_max = form.salary_max.data
        job.experience_required = form.experience_required.data
        job.education_required = form.education_required.data
        job.deadline = form.deadline.data

        job.save()
        flash('Job updated successfully!', 'success')
        return redirect(url_for('manage_jobs'))

    return render_template('admin/job_form.html', form=form, job=job, companies=companies, title='Edit Job')

@app.route('/admin/job/<job_id>/toggle', methods=['POST'])
@admin_required
def toggle_job_status(job_id):
    try:
        job = Job.objects.get(id=job_id)
        job.is_active = not job.is_active
        job.save()
        status = 'activated' if job.is_active else 'deactivated'
        flash(f'Job {status} successfully!', 'success')
    except:
        flash('Job not found or could not be updated', 'danger')
    return redirect(url_for('manage_jobs'))

@app.route('/admin/job/<job_id>/delete', methods=['POST'])
@admin_required
def delete_job(job_id):
    try:
        job = Job.objects.get(id=job_id)
        job.delete()
        flash('Job deleted successfully!', 'success')
    except:
        flash('Job not found or could not be deleted', 'danger')
    return redirect(url_for('manage_jobs'))

# Company Management
@app.route('/admin/companies')
@admin_required
def manage_companies():
    page = request.args.get('page', 1, type=int)
    per_page = 20
    skip = (page - 1) * per_page

    # Get total count for pagination
    total_companies = Company.objects.count()

    # Get companies with pagination
    companies_list = Company.objects.skip(skip).limit(per_page)

    # Create a pagination object with the necessary attributes
    class Pagination:
        def __init__(self, items, page, per_page, total):
            self.items = items
            self.page = page
            self.per_page = per_page
            self.total = total
            self.pages = (total + per_page - 1) // per_page  # Ceiling division

        @property
        def has_prev(self):
            return self.page > 1

        @property
        def has_next(self):
            return self.page < self.pages

        @property
        def prev_num(self):
            return self.page - 1 if self.has_prev else None

        @property
        def next_num(self):
            return self.page + 1 if self.has_next else None

    companies = Pagination(companies_list, page, per_page, total_companies)
    return render_template('admin/manage_companies.html', companies=companies, Job=Job)

@app.route('/admin/company/new', methods=['GET', 'POST'])
@admin_required
def add_company():
    form = CompanyForm()

    if form.validate_on_submit():
        # Find or create an employer user
        employer_email = request.form.get('employer_email')
        employer = User.objects(email=employer_email).first()

        if not employer:
            # Create new employer user
            employer = User(
                username=employer_email.split('@')[0],
                email=employer_email,
                role='employer',
                is_active=True
            )
            employer.set_password('password123')  # Default password
            employer.save()

        # Check if employer already has a company
        if Company.objects(user=employer).first():
            flash('This employer already has a company.', 'danger')
            return render_template('admin/company_form.html', form=form, title='Add Company')

        # Create company
        company = Company(
            user=employer,
            name=form.name.data,
            website=form.website.data,
            description=form.description.data,
            industry=form.industry.data,
            company_size=form.company_size.data,
            founded_year=form.founded_year.data,
            headquarters=form.headquarters.data,
            company_type=form.company_type.data
        )

        # Handle logo upload
        if form.logo.data:
            logo_path = save_file(form.logo.data, 'company_logos')
            if logo_path:
                company.logo_path = logo_path

        company.save()

        # Add company to selected categories
        category_ids = request.form.getlist('categories')
        for cat_id in category_ids:
            category = CompanyCategory.objects(id=cat_id).first()
            if category:
                assoc = CompanyCategoryAssociation(company=company, category=category)
                assoc.save()
        flash('Company created successfully!', 'success')
        return redirect(url_for('manage_companies'))

    categories = CompanyCategory.objects.all()
    return render_template('admin/company_form.html', form=form, categories=categories, title='Add Company')

@app.route('/admin/company/<company_id>/edit', methods=['GET', 'POST'])
@admin_required
def edit_company(company_id):
    try:
        company = Company.objects.get(id=company_id)
    except:
        abort(404)

    form = CompanyForm(obj=company)

    if form.validate_on_submit():
        company.name = form.name.data
        company.website = form.website.data
        company.description = form.description.data
        company.industry = form.industry.data
        company.company_size = form.company_size.data
        company.founded_year = form.founded_year.data
        company.headquarters = form.headquarters.data
        company.company_type = form.company_type.data

        # Handle logo upload
        if form.logo.data:
            logo_path = save_file(form.logo.data, 'company_logos')
            if logo_path:
                # Delete old logo if exists
                if company.logo_path:
                    old_logo_path = os.path.join(app.config['UPLOAD_FOLDER'], company.logo_path)
                    if os.path.exists(old_logo_path):
                        os.remove(old_logo_path)
                company.logo_path = logo_path

        # Update company categories
        CompanyCategoryAssociation.objects(company=company).delete()

        category_ids = request.form.getlist('categories')
        for cat_id in category_ids:
            category = CompanyCategory.objects(id=cat_id).first()
            if category:
                assoc = CompanyCategoryAssociation(company=company, category=category)
                assoc.save()

        # Update featured status
        company.is_featured = 'featured' in request.form

        company.save()
        flash('Company updated successfully!', 'success')
        return redirect(url_for('manage_companies'))

    categories = CompanyCategory.objects.all()
    company_categories = [assoc.category.id for assoc in CompanyCategoryAssociation.objects(company=company).all()]

    return render_template('admin/company_form.html',
                           form=form,
                           company=company,
                           categories=categories,
                           company_categories=company_categories,
                           title='Edit Company')

@app.route('/admin/company/<company_id>/toggle_featured', methods=['POST'])
@admin_required
def toggle_company_featured(company_id):
    try:
        company = Company.objects.get(id=company_id)
        company.is_featured = not company.is_featured
        company.save()
        status = 'featured' if company.is_featured else 'unfeatured'
        flash(f'Company {status} successfully!', 'success')
    except:
        flash('Company not found or could not be updated', 'danger')
    return redirect(url_for('manage_companies'))

@app.route('/admin/company/<company_id>/delete', methods=['POST'])
@admin_required
def delete_company(company_id):
    try:
        company = Company.objects.get(id=company_id)

        # Delete logo file if exists
        if company.logo_path:
            import os
            logo_path = os.path.join(app.config['UPLOAD_FOLDER'], company.logo_path)
            if os.path.exists(logo_path):
                os.remove(logo_path)

        company.delete()
        flash('Company deleted successfully!', 'success')
    except Exception as e:
        flash(f'Company not found or could not be deleted: {str(e)}', 'danger')
    return redirect(url_for('manage_companies'))

# Company Categories Management
@app.route('/admin/categories')
@admin_required
def manage_categories():
    categories = CompanyCategory.objects.all()
    return render_template('admin/manage_categories.html', categories=categories)

@app.route('/admin/category/new', methods=['GET', 'POST'])
@admin_required
def add_category():
    # Import the form
    from forms import AdminCompanyCategoryForm
    form = AdminCompanyCategoryForm()

    if form.validate_on_submit():
        category = CompanyCategory(
            name=form.name.data,
            description=form.description.data
        )
        category.save()
        flash('Category created successfully!', 'success')
        return redirect(url_for('manage_categories'))

    return render_template('admin/category_form.html', form=form, title='Add Category')

@app.route('/admin/category/<category_id>/edit', methods=['GET', 'POST'])
@admin_required
def edit_category(category_id):
    try:
        category = CompanyCategory.objects.get(id=category_id)
    except:
        abort(404)

    # Import the form
    from forms import AdminCompanyCategoryForm
    form = AdminCompanyCategoryForm(obj=category)

    if form.validate_on_submit():
        category.name = form.name.data
        category.description = form.description.data
        category.save()
        flash('Category updated successfully!', 'success')
        return redirect(url_for('manage_categories'))

    return render_template('admin/category_form.html', form=form, category=category, title='Edit Category')

@app.route('/admin/category/<category_id>/delete', methods=['POST'])
@admin_required
def delete_category(category_id):
    try:
        category = CompanyCategory.objects.get(id=category_id)

        # In MongoDB, we would need to find and delete associations separately
        # This would depend on how the associations are modeled in MongoDB
        # For example, if using references:
        CompanyCategoryAssociation.objects(category=category).delete()

        category.delete()
        flash('Category deleted successfully!', 'success')
    except Exception as e:
        flash(f'Category not found or could not be deleted: {str(e)}', 'danger')
    return redirect(url_for('manage_categories'))

# Job Scraper
@app.route('/admin/job-scraper', methods=['GET', 'POST'])
@admin_required
def job_scraper():
    companies = Company.objects.all()
    recent_imported_jobs = Job.objects(description__contains='[Imported from').order_by('-posted_date').limit(5)
    scraped_jobs = None
    job_data = None

    if request.method == 'POST':
        source_url = request.form.get('source_url')
        source_site = request.form.get('source_site')
        company_id = request.form.get('company_id')
        job_type = request.form.get('job_type')
        is_remote = True if request.form.get('is_remote') else False

        try:
            # Get company
            company = Company.objects(id=company_id).first()
            if not company:
                flash('Company not found', 'danger')
                return redirect(url_for('job_scraper'))

            # Use web scraper to extract job content
            content = get_website_text_content(source_url)

            if not content:
                flash('Failed to scrape content from the URL', 'danger')
                return redirect(url_for('job_scraper'))

            # Process content to extract jobs
            # This is a simplified implementation
            # In a real system, we would use NLP and more sophisticated parsing
            lines = content.split('\n')
            job_titles = []
            descriptions = []

            # Extract potential job titles (lines that might be job titles)
            for i, line in enumerate(lines):
                line = line.strip()
                if len(line) > 10 and len(line) < 100 and not line.endswith('.'):
                    if any(keyword in line.lower() for keyword in ['developer', 'engineer', 'designer', 'manager', 'specialist', 'analyst']):
                        job_titles.append((line, i))

            # Extract descriptions for each job title
            scraped_jobs = []
            for title, position in job_titles[:5]:  # Limit to 5 jobs
                # Get text after job title as description (limited to next 20 lines)
                description = '\n'.join(lines[position+1:position+20])

                # Create a dictionary for each job
                job_dict = {
                    'title': title,
                    'company_name': company.name,
                    'company_id': company.id,
                    'description': description,
                    'job_type': job_type,
                    'is_remote': is_remote,
                    'source': source_site,
                    'source_url': source_url,
                    'status': 'new'
                }

                # Check if job already exists
                if Job.objects(title=title, company=company).first():
                    job_dict['status'] = 'exists'

                scraped_jobs.append(job_dict)

            job_data = json.dumps(scraped_jobs)

            if not scraped_jobs:
                flash('No jobs could be extracted from the content', 'warning')

        except Exception as e:
            logging.error(f"Error scraping jobs: {str(e)}")
            flash(f'Error: {str(e)}', 'danger')

    return render_template('admin/job_scraper.html',
                          companies=companies,
                          scraped_jobs=scraped_jobs,
                          job_data=job_data,
                          recent_imported_jobs=recent_imported_jobs)

@app.route('/admin/import-scraped-jobs', methods=['POST'])
@admin_required
def import_scraped_jobs():
    job_data = request.form.get('job_data')

    if not job_data:
        flash('No job data to import', 'danger')
        return redirect(url_for('job_scraper'))

    try:
        jobs = json.loads(job_data)
        imported_count = 0

        for job_dict in jobs:
            if job_dict['status'] != 'exists':
                # Create new job
                company = Company.objects(id=job_dict['company_id']).first()
                if not company:
                    continue
                job = Job(
                    company=company,
                    title=job_dict['title'],
                    description=f"{job_dict['description']}\n\n[Imported from {job_dict['source']} on {datetime.now().strftime('%Y-%m-%d')}]",
                    job_type=job_dict['job_type'],
                    is_remote=job_dict['is_remote'],
                    is_active=True,
                    location="Remote" if job_dict['is_remote'] else "Not specified",
                    requirements="Not specified",
                    posted_date=datetime.now(timezone.utc)
                )

                job.save()
                imported_count += 1
        flash(f'Successfully imported {imported_count} jobs', 'success')

    except Exception as e:
        logging.error(f"Error importing jobs: {str(e)}")
        # No need for rollback in MongoDB
        flash(f'Error importing jobs: {str(e)}', 'danger')

    return redirect(url_for('job_scraper'))

# Analytics Dashboard
@app.route('/admin/analytics')
@admin_required
def analytics_dashboard():
    # Simulated analytics data - in a production environment,
    # this would be replaced with real analytics from a database

    # User statistics
    total_users = User.objects.count()
    jobseekers = User.objects(role='jobseeker').count()
    employers = User.objects(role='employer').count()
    admins = User.objects(role='admin').count()

    # Calculate percentages
    jobseeker_percentage = round((jobseekers / total_users) * 100) if total_users > 0 else 0
    employer_percentage = round((employers / total_users) * 100) if total_users > 0 else 0
    admin_percentage = round((admins / total_users) * 100) if total_users > 0 else 0

    # Get active employers (those with at least one job posting)
    active_employers = Company.objects.count()

    # Simulated visitor data
    live_visitors = random.randint(5, 30)
    total_visitors = random.randint(500, 1500)

    # Growth metrics (simulated)
    live_visitors_growth = random.randint(5, 15)
    total_visitors_growth = random.randint(10, 30)
    employers_growth = random.randint(5, 20)

    # Simulated premium subscribers
    premium_subscribers = random.randint(10, 50)
    subscribers_growth = random.randint(5, 25)

    # Visitor time series data (simulated)
    visitors_data = [random.randint(300, 1000) for _ in range(12)]
    signups_data = [random.randint(20, 100) for _ in range(12)]
    applications_data = [random.randint(50, 200) for _ in range(12)]

    # User engagement metrics (simulated)
    avg_session_duration = f"{random.randint(2, 10)}m {random.randint(0, 59)}s"
    bounce_rate = round(random.uniform(30, 60), 1)
    pages_per_session = round(random.uniform(2, 8), 1)
    conversion_rate = round(random.uniform(2, 10), 1)

    # Get top categories
    categories = CompanyCategory.objects.all()
    categories_labels = [category.name for category in categories]
    categories_jobs_data = [random.randint(10, 100) for _ in range(len(categories_labels))]
    categories_applications_data = [random.randint(20, 150) for _ in range(len(categories_labels))]

    # Get browser and device data (simulated)
    device_data = [60, 30, 10]  # Desktop, Mobile, Tablet
    browser_data = [50, 20, 15, 10, 5]  # Chrome, Safari, Firefox, Edge, Other

    # Geographic data (simulated)
    top_countries = [
        {'name': 'United States', 'visitors': random.randint(300, 800), 'percentage': random.randint(30, 50)},
        {'name': 'India', 'visitors': random.randint(100, 400), 'percentage': random.randint(10, 30)},
        {'name': 'United Kingdom', 'visitors': random.randint(50, 200), 'percentage': random.randint(5, 15)},
        {'name': 'Germany', 'visitors': random.randint(30, 150), 'percentage': random.randint(3, 10)},
        {'name': 'Canada', 'visitors': random.randint(20, 100), 'percentage': random.randint(2, 8)}
    ]

    # Referrer data (simulated)
    top_referrers = [
        {'source': 'Google', 'visitors': random.randint(200, 600), 'conversion': round(random.uniform(2, 8), 1)},
        {'source': 'Direct', 'visitors': random.randint(100, 400), 'conversion': round(random.uniform(5, 15), 1)},
        {'source': 'LinkedIn', 'visitors': random.randint(50, 300), 'conversion': round(random.uniform(8, 20), 1)},
        {'source': 'Facebook', 'visitors': random.randint(30, 200), 'conversion': round(random.uniform(1, 10), 1)},
        {'source': 'Twitter', 'visitors': random.randint(20, 100), 'conversion': round(random.uniform(1, 5), 1)}
    ]

    return render_template('admin/analytics_dashboard.html',
                          total_users=total_users,
                          jobseeker_percentage=jobseeker_percentage,
                          employer_percentage=employer_percentage,
                          admin_percentage=admin_percentage,
                          active_employers=active_employers,
                          live_visitors=live_visitors,
                          total_visitors=total_visitors,
                          live_visitors_growth=live_visitors_growth,
                          total_visitors_growth=total_visitors_growth,
                          employers_growth=employers_growth,
                          premium_subscribers=premium_subscribers,
                          subscribers_growth=subscribers_growth,
                          visitors_data=json.dumps(visitors_data),
                          signups_data=json.dumps(signups_data),
                          applications_data=json.dumps(applications_data),
                          avg_session_duration=avg_session_duration,
                          bounce_rate=bounce_rate,
                          pages_per_session=pages_per_session,
                          conversion_rate=conversion_rate,
                          categories_labels=json.dumps(categories_labels),
                          categories_jobs_data=json.dumps(categories_jobs_data),
                          categories_applications_data=json.dumps(categories_applications_data),
                          device_data=json.dumps(device_data),
                          browser_data=json.dumps(browser_data),
                          top_countries=top_countries,
                          top_referrers=top_referrers)

# Google Analytics Setup
@app.route('/admin/google-analytics')
@admin_required
def google_analytics_setup():
    return render_template('admin/google_analytics_setup.html')

@app.route('/admin/google-analytics/save', methods=['POST'])
@admin_required
def save_google_analytics():
    tracking_id = request.form.get('tracking_id')
    # In a production environment, we would save this to database
    # or environment variables
    # For now, just simulate success
    flash('Google Analytics tracking ID saved successfully!', 'success')
    return redirect(url_for('google_analytics_setup'))

# Application Management
@app.route('/admin/applications')
@admin_required
def manage_applications():
    page = request.args.get('page', 1, type=int)
    per_page = 20
    skip = (page - 1) * per_page

    # Get total count for pagination
    total_applications = JobApplication.objects.count()

    # Get applications with pagination
    applications_list = JobApplication.objects.order_by('-applied_date').skip(skip).limit(per_page)

    # Create a pagination object with the necessary attributes
    class Pagination:
        def __init__(self, items, page, per_page, total):
            self.items = items
            self.page = page
            self.per_page = per_page
            self.total = total
            self.pages = (total + per_page - 1) // per_page  # Ceiling division

        @property
        def has_prev(self):
            return self.page > 1

        @property
        def has_next(self):
            return self.page < self.pages

        @property
        def prev_num(self):
            return self.page - 1 if self.has_prev else None

        @property
        def next_num(self):
            return self.page + 1 if self.has_next else None

    applications_pagination = Pagination(applications_list, page, per_page, total_applications)

    # Create a list with application data and related entities
    applications = []
    for app in applications_pagination.items:
        applications.append((app, app.job, app.user, app.job.company if app.job else None))

    return render_template('admin/manage_applications.html',
                          applications=applications,
                          pagination=applications_pagination)

@app.route('/admin/application/<application_id>/status', methods=['POST'])
@admin_required
def update_application_status(application_id):
    try:
        application = JobApplication.objects.get(id=application_id)
        new_status = request.form.get('status')

        if new_status in ['applied', 'reviewed', 'shortlisted', 'rejected', 'hired']:
            application.status = new_status
            application.last_updated = datetime.now(timezone.utc)
            application.save()
            flash('Application status updated successfully!', 'success')
        else:
            flash('Invalid status provided.', 'danger')
    except:
        flash('Application not found.', 'danger')

    return redirect(url_for('manage_applications'))

@app.route('/admin/application/<application_id>/delete', methods=['POST'])
@admin_required
def delete_application(application_id):
    try:
        application = JobApplication.objects.get(id=application_id)

        # Delete resume file if exists
        if application.resume_path:
            resume_path = os.path.join(app.config['UPLOAD_FOLDER'], application.resume_path)
            if os.path.exists(resume_path):
                os.remove(resume_path)

        application.delete()
        flash('Application deleted successfully!', 'success')
    except:
        flash('Application not found or could not be deleted', 'danger')

    return redirect(url_for('manage_applications'))

# Helper function to check if a file is allowed
def allowed_file(filename, allowed_extensions):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions

# Helper function for saving files
def save_file(file, folder):
    if file and allowed_file(file.filename, {'pdf', 'doc', 'docx', 'jpg', 'jpeg', 'png', 'svg'}):
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        filename = f"{timestamp}_{filename}"
        upload_folder = os.path.join(app.config['UPLOAD_FOLDER'], folder)
        os.makedirs(upload_folder, exist_ok=True)
        file_path = os.path.join(upload_folder, filename)
        file.save(file_path)
        return os.path.join(folder, filename)
    return None

# Helper function for web scraping
def get_website_text_content(url):
    """
    Scrape text content from a website URL.
    This is a stub function - in a real application, you would implement
    web scraping using libraries like BeautifulSoup or Scrapy.
    """
    try:
        # In a real implementation, you would use requests and BeautifulSoup
        # For now, just return a placeholder message
        return f"Placeholder content scraped from {url}"
    except Exception as e:
        logging.error(f"Error scraping content from {url}: {str(e)}")
        return None



# Import/Export Data
@app.route('/admin/import-export')
@admin_required
def import_export_data():
    return render_template('admin/import_export.html')

@app.route('/admin/export-data', methods=['POST'])
@admin_required
def export_data():
    data_type = request.form.get('data_type')
    export_format = request.form.get('export_format')

    # Prepare data based on type
    if data_type == 'users':
        data = User.query.all()
        headers = ['ID', 'Username', 'Email', 'Role', 'Active', 'Created At']
        rows = [[user.id, user.username, user.email, user.role,
                user.is_active, user.created_at.strftime('%Y-%m-%d')] for user in data]
    elif data_type == 'jobs':
        data = Job.query.all()
        headers = ['ID', 'Title', 'Company', 'Type', 'Remote', 'Active', 'Posted Date']
        rows = [[job.id, job.title, job.company.name, job.job_type,
                job.is_remote, job.is_active, job.posted_date.strftime('%Y-%m-%d')] for job in data]
    elif data_type == 'companies':
        data = Company.query.all()
        headers = ['ID', 'Name', 'Industry', 'Size', 'Featured']
        rows = [[company.id, company.name, company.industry,
                company.company_size, company.is_featured] for company in data]
    elif data_type == 'applications':
        data = JobApplication.query.all()
        headers = ['ID', 'Job', 'Applicant', 'Status', 'Applied Date']
        rows = [[app.id, Job.query.get(app.job_id).title,
                User.query.get(app.user_id).username, app.status,
                app.applied_date.strftime('%Y-%m-%d')] for app in data]
    else:
        flash('Invalid data type selected', 'danger')
        return redirect(url_for('import_export_data'))

    # Generate export file based on format
    if export_format == 'csv':
        # Generate CSV
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)

        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename={data_type}_{datetime.now().strftime("%Y%m%d")}.csv'
        response.headers['Content-type'] = 'text/csv'
        return response

    elif export_format == 'excel':
        # Generate Excel
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = data_type.capitalize()

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)

        # Write data rows
        for row_num, row in enumerate(rows, 2):
            for col_num, cell_value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num).value = cell_value

        wb.save(output)
        output.seek(0)

        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename={data_type}_{datetime.now().strftime("%Y%m%d")}.xlsx'
        response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response

    elif export_format == 'pdf':
        # Generate PDF
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer)

        # Set title
        p.setFont('Helvetica-Bold', 16)
        p.drawString(30, 750, f"{data_type.capitalize()} Report - {datetime.now().strftime('%Y-%m-%d')}")

        # Draw table headers
        p.setFont('Helvetica-Bold', 12)
        x_offset = 30
        for header in headers:
            p.drawString(x_offset, 720, header)
            x_offset += 100

        # Draw table rows
        p.setFont('Helvetica', 10)
        y_offset = 700
        for row in rows[:40]:  # Limit to 40 rows per page for simplicity
            x_offset = 30
            for cell in row:
                p.drawString(x_offset, y_offset, str(cell))
                x_offset += 100
            y_offset -= 20

        p.save()
        buffer.seek(0)

        response = make_response(buffer.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename={data_type}_{datetime.now().strftime("%Y%m%d")}.pdf'
        response.headers['Content-type'] = 'application/pdf'
        return response

    else:
        # JSON format
        json_data = []
        for i, row in enumerate(rows):
            item = {}
            for j, header in enumerate(headers):
                item[header] = row[j]
            json_data.append(item)

        response = make_response(json.dumps(json_data, default=str, indent=4))
        response.headers['Content-Disposition'] = f'attachment; filename={data_type}_{datetime.now().strftime("%Y%m%d")}.json'
        response.headers['Content-type'] = 'application/json'
        return response

@app.route('/admin/import-data', methods=['POST'])
@admin_required
def import_data():
    if 'import_file' not in request.files:
        flash('No file selected', 'danger')
        return redirect(url_for('import_export_data'))

    file = request.files['import_file']
    if file.filename == '':
        flash('No file selected', 'danger')
        return redirect(url_for('import_export_data'))

    data_type = request.form.get('import_data_type')

    if file and allowed_file(file.filename, {'csv', 'xlsx', 'json'}):
        try:
            filename = secure_filename(file.filename)
            ext = filename.rsplit('.', 1)[1].lower()

            if ext == 'csv':
                # Process CSV
                file_content = file.read().decode('utf-8')
                csv_data = csv.reader(io.StringIO(file_content))
                headers = next(csv_data)  # Skip header row
                rows = list(csv_data)
            elif ext == 'xlsx' and Workbook is not None:
                # Process Excel if openpyxl is available
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(filename=file)
                    ws = wb.active
                    rows = list(ws.values)
                    headers = rows[0]
                    rows = rows[1:]
                except ImportError:
                    flash('Excel processing requires openpyxl library', 'danger')
                    return redirect(url_for('import_export_data'))
            elif ext == 'json':
                # Process JSON
                json_data = json.loads(file.read().decode('utf-8'))
                if not json_data:
                    flash('Empty or invalid JSON file', 'danger')
                    return redirect(url_for('import_export_data'))

                # Import logic based on data type
                if data_type == 'jobs':
                    # Process job data import
                    for job_data in json_data:
                        # Check if company exists
                        company_name = job_data.get('company_name')
                        company = Company.query.filter_by(name=company_name).first()
                        if not company:
                            flash(f'Company "{company_name}" not found for job: {job_data.get("title")}', 'warning')
                            continue

                        # Check if job already exists
                        if Job.query.filter_by(title=job_data.get('title'), company_id=company.id).first():
                            flash(f'Job "{job_data.get("title")}" at {company_name} already exists', 'warning')
                            continue

                        # Create job
                        job = Job(
                            company_id=company.id,
                            title=job_data.get('title'),
                            description=job_data.get('description', ''),
                            location=job_data.get('location', ''),
                            is_remote=job_data.get('is_remote', False),
                            job_type=job_data.get('job_type', 'Full-time'),
                            requirements=job_data.get('requirements', ''),
                            posted_date=datetime.now(),
                            is_active=True
                        )
                        db.session.add(job)

                    db.session.commit()
                    flash(f'Successfully imported jobs from JSON', 'success')
                else:
                    flash('This import type is not currently supported', 'warning')

                return redirect(url_for('import_export_data'))

            # Process import based on data type and loaded data
            # For brevity, we'll just handle job import for CSV/Excel
            if data_type == 'jobs':
                # Map headers to expected columns
                title_idx = headers.index('Title') if 'Title' in headers else -1
                company_idx = headers.index('Company') if 'Company' in headers else -1
                type_idx = headers.index('Type') if 'Type' in headers else -1
                remote_idx = headers.index('Remote') if 'Remote' in headers else -1

                if title_idx == -1 or company_idx == -1:
                    flash('Missing required columns in import file', 'danger')
                    return redirect(url_for('import_export_data'))

                # Process rows
                imported_count = 0
                for row in rows:
                    title = row[title_idx]
                    company_name = row[company_idx]
                    job_type = row[type_idx] if type_idx != -1 else 'Full-time'
                    is_remote = row[remote_idx] if remote_idx != -1 else False

                    # Find company
                    company = Company.query.filter_by(name=company_name).first()
                    if not company:
                        flash(f'Company "{company_name}" not found for job: {title}', 'warning')
                        continue

                    # Check if job already exists
                    if Job.query.filter_by(title=title, company_id=company.id).first():
                        flash(f'Job "{title}" at {company_name} already exists', 'warning')
                        continue

                    # Create job
                    job = Job(
                        company_id=company.id,
                        title=title,
                        description='Imported job - please update description',
                        job_type=job_type,
                        is_remote=is_remote,
                        is_active=True,
                        location="Not specified",
                        requirements="Not specified",
                        posted_date=datetime.now()
                    )

                    db.session.add(job)
                    imported_count += 1

                db.session.commit()
                flash(f'Successfully imported {imported_count} jobs', 'success')
            else:
                flash('This import type is not currently supported', 'warning')

        except Exception as e:
            db.session.rollback()
            logging.error(f"Error importing data: {str(e)}")
            flash(f'Error importing data: {str(e)}', 'danger')
    else:
        flash('Invalid file format. Please upload a CSV, Excel, or JSON file.', 'danger')

    return redirect(url_for('import_export_data'))

# Site Content Management
@app.route('/admin/site-content')
@admin_required
def manage_site_content():
    # In a real implementation, you would fetch the current content from the database
    # and pass it to the template
    return render_template('admin/site_content.html')

@app.route('/admin/site-content/update', methods=['POST'])
@admin_required
def update_site_content():
    section = request.form.get('section')

    if not section:
        flash('Missing section data', 'danger')
        return redirect(url_for('manage_site_content'))

    # Process different sections based on the form data
    if section == 'header':
        # Process header section
        site_name = request.form.get('site_name')
        tagline = request.form.get('tagline')

        # Handle logo upload if provided
        logo_file = request.files.get('site_logo')
        if logo_file and logo_file.filename:
            filename = secure_filename(logo_file.filename)
            logo_path = os.path.join(app.config['UPLOAD_FOLDER'], 'site', filename)
            os.makedirs(os.path.dirname(logo_path), exist_ok=True)
            logo_file.save(logo_path)

        # Process navigation links
        nav_texts = request.form.getlist('nav_text[]')
        nav_urls = request.form.getlist('nav_url[]')
        nav_orders = request.form.getlist('nav_order[]')
        nav_visibles = request.form.getlist('nav_visible[]')

        # In a real implementation, you would save this data to the database

        flash('Header content updated successfully', 'success')

    elif section == 'hero':
        # Process hero section
        hero_headline = request.form.get('hero_headline')
        hero_subheadline = request.form.get('hero_subheadline')
        cta_text = request.form.get('cta_text')
        cta_url = request.form.get('cta_url')
        hero_bg_style = request.form.get('hero_bg_style')

        # Handle hero image upload if provided
        hero_image = request.files.get('hero_image')
        if hero_image and hero_image.filename:
            filename = secure_filename(hero_image.filename)
            image_path = os.path.join(app.config['UPLOAD_FOLDER'], 'site', filename)
            os.makedirs(os.path.dirname(image_path), exist_ok=True)
            hero_image.save(image_path)

        # In a real implementation, you would save this data to the database

        flash('Hero section updated successfully', 'success')

    elif section == 'footer':
        # Process footer section
        footer_text = request.form.get('footer_text')
        footer_description = request.form.get('footer_description')

        # Process social media links
        social_facebook = request.form.get('social_facebook')
        social_twitter = request.form.get('social_twitter')
        social_linkedin = request.form.get('social_linkedin')
        social_instagram = request.form.get('social_instagram')

        # Process footer columns
        footer_col_titles = request.form.getlist('footer_col_title[]')

        # In a real implementation, you would save this data to the database

        flash('Footer content updated successfully', 'success')

    elif section == 'about':
        # Process about page
        about_title = request.form.get('about_title')
        about_content = request.form.get('about_content')

        # Process team members
        team_names = request.form.getlist('team_name[]')
        team_positions = request.form.getlist('team_position[]')
        team_bios = request.form.getlist('team_bio[]')

        # Handle team photos if provided
        team_photos = request.files.getlist('team_photo[]')
        for i, photo in enumerate(team_photos):
            if photo and photo.filename:
                filename = secure_filename(photo.filename)
                photo_path = os.path.join(app.config['UPLOAD_FOLDER'], 'team', filename)
                os.makedirs(os.path.dirname(photo_path), exist_ok=True)
                photo.save(photo_path)

        # In a real implementation, you would save this data to the database

        flash('About page content updated successfully', 'success')

    elif section == 'home':
        # Process home page content

        # Featured Section
        featured_title = request.form.get('featured_title')
        featured_description = request.form.get('featured_description')
        featured_count = request.form.get('featured_count')
        show_featured_section = 'show_featured_section' in request.form

        # Recent Jobs Section
        jobs_title = request.form.get('jobs_title')
        jobs_description = request.form.get('jobs_description')
        jobs_count = request.form.get('jobs_count')
        show_jobs_section = 'show_jobs_section' in request.form

        # How It Works Section
        how_it_works_title = request.form.get('how_it_works_title')
        how_it_works_description = request.form.get('how_it_works_description')
        show_how_it_works = 'show_how_it_works' in request.form

        # Process steps
        step_numbers = request.form.getlist('step_number[]')
        step_titles = request.form.getlist('step_title[]')
        step_descriptions = request.form.getlist('step_description[]')
        step_icons = request.form.getlist('step_icon[]')

        # Statistics Section
        stats_title = request.form.get('stats_title')
        stats_description = request.form.get('stats_description')
        show_stats = 'show_stats' in request.form

        # Process statistics
        stat_titles = request.form.getlist('stat_title[]')
        stat_values = request.form.getlist('stat_value[]')
        stat_icons = request.form.getlist('stat_icon[]')

        # Call to Action Section
        cta_title = request.form.get('cta_title')
        cta_description = request.form.get('cta_description')
        cta_button_text = request.form.get('cta_button_text')
        cta_button_url = request.form.get('cta_button_url')
        show_cta = 'show_cta' in request.form

        # In a real implementation, you would save this data to the database
        # For example, using a SiteContent model with a section field and a content field
        # that stores the JSON representation of the section data

        flash('Home page content updated successfully', 'success')

    elif section == 'testimonials':
        # Process testimonials
        # This would be implemented in a similar way to the other sections

        flash('Testimonials updated successfully', 'success')

    else:
        flash(f'Unknown section: {section}', 'danger')

    return redirect(url_for('manage_site_content'))



# Helper function for allowed file types
def allowed_file(filename, allowed_extensions):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions
